"""Rebuild the archived July STEO data snapshot."""

from argparse import ArgumentParser
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
from shutil import copyfileobj
from tempfile import TemporaryDirectory
from urllib.request import Request, urlopen

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_MANIFEST = PROJECT_ROOT / "data" / "eia_steo_vintage_manifest.csv"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "eia_steo_july_vintages.csv"

# The first item gives the output column, followed by the workbook sheet and
# the EIA series code. The code is used instead of a fixed row number because
# some table layouts changed in 2024.
SERIES = {
    "price": ("2tab", "NGHHUUS"),
    "storage": ("5atab", "NGWGPUS"),
    "hdd": ("9ctab", "ZWHDPUS"),
    "hdd_normal": ("9ctab", "ZWHD_US_10YR"),
    "cdd": ("9ctab", "ZWCDPUS"),
    "cdd_normal": ("9ctab", "ZWCD_US_10YR"),
}


def file_sha256(path):
    """Return the SHA-256 checksum of one file."""
    checksum = sha256()
    with path.open("rb") as file:
        for block in iter(lambda: file.read(1024 * 1024), b""):
            checksum.update(block)
    return checksum.hexdigest()


def download_workbook(url, destination):
    """Download one archived workbook from its official EIA URL."""
    request = Request(url, headers={"User-Agent": "GasStorageValuation/1.0"})
    with urlopen(request, timeout=120) as response, destination.open("wb") as file:
        copyfileobj(response, file)


def parse_month(value):
    """Convert the YYYYMM values used by the Dates sheet to month starts."""
    if isinstance(value, (datetime, date)):
        return pd.Timestamp(value.year, value.month, 1)

    try:
        text = str(int(value))
    except (TypeError, ValueError) as error:
        raise ValueError(f"cannot read workbook month {value!r}") from error

    if len(text) != 6:
        raise ValueError(f"cannot read workbook month {value!r}")
    return pd.Timestamp(int(text[:4]), int(text[4:]), 1)


def find_series_row(sheet, series_code):
    """Find the unique row containing an EIA series code."""
    matches = []
    first_column = sheet.iter_rows(min_col=1, max_col=1, values_only=True)
    for row_number, (value,) in enumerate(first_column, start=1):
        if value == series_code:
            matches.append(row_number)

    if len(matches) != 1:
        raise ValueError(
            f"expected one {series_code} row in {sheet.title}, found {len(matches)}"
        )
    return matches[0]


def extract_workbook(path, manifest_row):
    """Extract the required monthly series from one STEO workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        dates_sheet = workbook["Dates"]
        expected_dates = pd.date_range(
            manifest_row.first_month,
            manifest_row.last_month,
            freq="MS",
        )

        # Monthly dates are stored from column C across row 11.
        raw_dates = next(
            dates_sheet.iter_rows(
                min_row=11,
                max_row=11,
                min_col=3,
                max_col=2 + len(expected_dates),
                values_only=True,
            )
        )
        workbook_dates = pd.DatetimeIndex(parse_month(value) for value in raw_dates)
        if not workbook_dates.equals(expected_dates):
            raise ValueError(f"date coverage does not match the manifest for {path.name}")

        # Dates!D7 records the final historical or estimated month in the file.
        last_historical_month = parse_month(dates_sheet["D7"].value)
        if last_historical_month != manifest_row.last_historical_month:
            raise ValueError(
                f"historical cutoff does not match the manifest for {path.name}"
            )

        extracted = {
            "vintage_year": [int(manifest_row.vintage_year)] * len(workbook_dates),
            "release_date": [manifest_row.release_date] * len(workbook_dates),
            "last_historical_month": [last_historical_month] * len(workbook_dates),
            "date": workbook_dates,
            "status": [
                "history_or_estimate" if month <= last_historical_month else "forecast"
                for month in workbook_dates
            ],
        }

        for output_column, (sheet_name, series_code) in SERIES.items():
            sheet = workbook[sheet_name]
            row_number = find_series_row(sheet, series_code)
            values = next(
                sheet.iter_rows(
                    min_row=row_number,
                    max_row=row_number,
                    min_col=3,
                    max_col=2 + len(workbook_dates),
                    values_only=True,
                )
            )
            if any(value is None for value in values):
                raise ValueError(f"{series_code} contains missing values in {path.name}")
            extracted[output_column] = values

        return pd.DataFrame(extracted)
    finally:
        workbook.close()


def read_manifest(path):
    """Read and check the pinned workbook manifest."""
    manifest = pd.read_csv(
        path,
        parse_dates=[
            "release_date",
            "retrieved_date",
            "last_historical_month",
            "first_month",
            "last_month",
        ],
    ).sort_values("vintage_year")

    if manifest["vintage_year"].duplicated().any():
        raise ValueError("the manifest contains duplicate vintage years")
    if manifest["workbook_filename"].duplicated().any():
        raise ValueError("the manifest contains duplicate workbook filenames")
    return manifest


def validate_snapshot(data, manifest):
    """Run the main completeness checks before writing the CSV."""
    counts = data.groupby("vintage_year")["date"].size()
    if not counts.eq(72).all() or len(counts) != len(manifest):
        raise ValueError("each vintage must contain exactly 72 monthly rows")
    if data.duplicated(["vintage_year", "date"]).any():
        raise ValueError("the extracted data contain duplicate vintage-month rows")

    value_columns = list(SERIES)
    if data[value_columns].isna().any().any():
        raise ValueError("the extracted data contain missing series values")
    if not (data[["price", "storage"]] > 0).all().all():
        raise ValueError("price and storage values must be positive")
    if not (data[["hdd", "hdd_normal", "cdd", "cdd_normal"]] >= 0).all().all():
        raise ValueError("degree-day values cannot be negative")


def rebuild_snapshot(manifest_path, output_path):
    """Download, verify and combine every workbook in the manifest."""
    manifest = read_manifest(manifest_path)
    vintage_frames = []

    with TemporaryDirectory(prefix="eia_steo_") as temporary_directory:
        download_directory = Path(temporary_directory)

        for manifest_row in manifest.itertuples(index=False):
            workbook_path = download_directory / manifest_row.workbook_filename
            print(f"Downloading {manifest_row.workbook_filename}")
            download_workbook(manifest_row.source_url, workbook_path)

            actual_checksum = file_sha256(workbook_path)
            if actual_checksum != manifest_row.sha256:
                raise ValueError(
                    f"checksum mismatch for {manifest_row.workbook_filename}: "
                    f"expected {manifest_row.sha256}, found {actual_checksum}"
                )

            vintage_frames.append(extract_workbook(workbook_path, manifest_row))

    snapshot = pd.concat(vintage_frames, ignore_index=True)
    validate_snapshot(snapshot, manifest)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    snapshot.to_csv(output_path, index=False, date_format="%Y-%m-%d")
    print(f"Wrote {len(snapshot):,} rows to {output_path}")
    print(f"SHA-256: {file_sha256(output_path)}")


def parse_arguments():
    parser = ArgumentParser(description=__doc__)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_arguments()
    rebuild_snapshot(arguments.manifest, arguments.output)
